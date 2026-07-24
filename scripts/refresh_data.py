#!/usr/bin/env python3
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data" / "ships.csv"
JSON_PATH = ROOT / "data" / "dashboard-data.json"
CAPTURE_DIR = Path(
    "/Users/petersargent/Library/CloudStorage/"
    "GoogleDrive-petersargent@seascout.org/"
    "Shared drives/NSSC Membership Spreadsheet Capture"
)
CAPTURE_PATTERN = "SeaScoutShips-*.xlsx"
SOURCE_SHEET_ID = "1Y5DqDHs3qyH8Kx9z4jPSRIg0Fq0Ju-sDCxfVhRgx_dM"
SOURCE_URL = f"https://docs.google.com/spreadsheets/d/{SOURCE_SHEET_ID}/edit"

MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def clean_key(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%b-%y")
    if isinstance(value, date):
        return value.strftime("%b-%y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def to_int(value):
    if isinstance(value, bool) or value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = clean_text(value).replace(",", "")
    match = re.search(r"-?\d+", text)
    return int(match.group(0)) if match else 0


def to_delta(value):
    if isinstance(value, bool) or value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = clean_text(value)
    if not text:
        return 0
    sign = -1 if "▼" in text or text.startswith("-") else 1
    return sign * to_int(text)


def to_percent(value):
    if isinstance(value, bool) or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value) * 100, 2)
    text = clean_text(value).replace("%", "")
    try:
        percent = float(text)
        return round(percent * 100, 2) if 0 < abs(percent) <= 1 else percent
    except ValueError:
        return 0.0


def renewal_date(value):
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    text = clean_text(value).lower()
    match = re.match(r"([a-z]{3})-(\d{2})", text)
    if not match:
        return None
    month = MONTHS.get(match.group(1))
    if not month:
        return None
    year = 2000 + int(match.group(2))
    return date(year, month, 1)


def yes(value):
    if isinstance(value, bool):
        return value
    return clean_text(value).lower() in {"yes", "true", "1"}


def rate(part, whole):
    return round((part / whole) * 100, 1) if whole else 0


def top_counts(rows, key, limit=10):
    counts = Counter(row.get(key) or "Unassigned" for row in rows)
    return [{"name": name, "count": count} for name, count in counts.most_common(limit)]


def bucket_metric(value):
    metric = to_int(value)
    if metric <= 1:
        return "0-1"
    if metric == 2:
        return "2"
    if metric == 3:
        return "3"
    return "4-5"


def latest_capture():
    captures = sorted(
        CAPTURE_DIR.glob(CAPTURE_PATTERN),
        key=lambda path: (path.stem, path.stat().st_mtime),
        reverse=True,
    )
    if not captures:
        raise FileNotFoundError(f"No captured spreadsheets found in {CAPTURE_DIR}")
    return captures[0]


def format_csv_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%b-%y")
    if isinstance(value, date):
        return value.strftime("%b-%y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def write_csv_snapshot(rows, headers):
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: format_csv_value(row.get(header)) for header in headers})


def load_rows(capture_path):
    workbook = load_workbook(capture_path, read_only=True, data_only=True)
    if "Ships" not in workbook.sheetnames:
        raise ValueError(f"{capture_path} does not contain a Ships worksheet")

    worksheet = workbook["Ships"]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [clean_key(value) for value in next(iterator)]
    rows = []
    for raw in iterator:
        row = {header: value for header, value in zip(headers, raw)}
        if row.get("UnitID"):
            rows.append(row)

    write_csv_snapshot(rows, headers)
    return rows


def build_payload(rows, capture_path):
    generated = datetime.now().astimezone()
    today = generated.date()

    ships = []
    for row in rows:
        renewal = renewal_date(row.get("Renewal"))
        youth = to_int(row.get("Total Youth"))
        adult = to_int(row.get("Total Adult"))
        metric = to_int(row.get("Unit Metric"))
        ships.append(
            {
                "unitId": clean_text(row.get("UnitID")),
                "cst": to_int(row.get("CST")),
                "shipNumber": clean_text(row.get("Ship #")),
                "shipName": clean_text(row.get("Ship Name if tracking")),
                "council": clean_text(row.get("Council Name")),
                "district": clean_text(row.get("District")),
                "owner": clean_text(row.get("Owner")) or "Unassigned",
                "auxDistrict": clean_text(row.get("CG Aux Districts")) or "Unassigned",
                "focus": clean_text(row.get("Ship Focus")) or "Unspecified",
                "status": clean_text(row.get("Status")) or "Established",
                "youth": youth,
                "adult": adult,
                "youthYoY": to_delta(row.get("Total Youth YOY")),
                "primaryYouth": to_int(row.get("Primary Youth")),
                "primaryYoY": to_delta(row.get("Primary YOY")),
                "advancementRate": to_percent(row.get("%Advancement")),
                "retentionRate": to_percent(row.get("Retention")),
                "unitMetric": metric,
                "metricBucket": bucket_metric(row.get("Unit Metric")),
                "renewal": clean_text(row.get("Renewal")),
                "renewalDate": renewal.isoformat() if renewal else "",
                "skipperTrained": yes(row.get("Skipper Trained")),
                "committeeTrained": yes(row.get("CC Trained")),
                "sizeHealthy": yes(row.get("Size")),
                "growthHealthy": yes(row.get("Growth")),
                "advancementHealthy": yes(row.get("20% Advance")),
                "outdoorHealthy": yes(row.get("Outdoor")),
                "charterOrganization": clean_text(row.get("Charter Organization")),
                "skipperName": clean_text(row.get("Skipper Name")),
            }
        )

    total = len(ships)
    metric_counter = Counter(ship["unitMetric"] for ship in ships)
    renewal_counter = Counter(ship["renewal"] or "Unscheduled" for ship in ships)
    focus_counter = Counter(ship["focus"] for ship in ships)
    aux_counter = Counter(ship["auxDistrict"] for ship in ships)
    owner_counter = Counter(ship["owner"] for ship in ships)
    council_youth = defaultdict(int)
    council_ships = defaultdict(int)

    for ship in ships:
        council_youth[ship["council"]] += ship["youth"]
        council_ships[ship["council"]] += 1

    upcoming = []
    for ship in ships:
        if not ship["renewalDate"]:
            continue
        renews = date.fromisoformat(ship["renewalDate"])
        if 0 <= (renews - today).days <= 120:
            upcoming.append(ship)

    payload = {
        "meta": {
            "sourceUrl": SOURCE_URL,
            "sourceCapture": str(capture_path),
            "sourceCaptureName": capture_path.name,
            "generatedAt": generated.isoformat(timespec="seconds"),
            "rowCount": total,
        },
        "summary": {
            "ships": total,
            "newShips": sum(1 for ship in ships if ship["status"] == "New"),
            "youth": sum(ship["youth"] for ship in ships),
            "adults": sum(ship["adult"] for ship in ships),
            "youthYoY": sum(ship["youthYoY"] for ship in ships),
            "primaryYouth": sum(ship["primaryYouth"] for ship in ships),
            "primaryYoY": sum(ship["primaryYoY"] for ship in ships),
            "avgUnitMetric": round(
                sum(ship["unitMetric"] for ship in ships) / total, 2
            )
            if total
            else 0,
            "skipperTrainingRate": rate(
                sum(1 for ship in ships if ship["skipperTrained"]), total
            ),
            "committeeTrainingRate": rate(
                sum(1 for ship in ships if ship["committeeTrained"]), total
            ),
            "advancementHealthyRate": rate(
                sum(1 for ship in ships if ship["advancementHealthy"]), total
            ),
            "outdoorHealthyRate": rate(
                sum(1 for ship in ships if ship["outdoorHealthy"]), total
            ),
            "upcomingRenewals": len(upcoming),
        },
        "breakdowns": {
            "unitMetric": [
                {"name": str(name), "count": count}
                for name, count in sorted(metric_counter.items())
            ],
            "renewal": [
                {"name": name, "count": count}
                for name, count in sorted(renewal_counter.items())
            ],
            "focus": [{"name": name, "count": count} for name, count in focus_counter.most_common()],
            "auxDistrict": [
                {"name": name, "count": count} for name, count in aux_counter.most_common()
            ],
            "owner": [
                {"name": name, "count": count} for name, count in owner_counter.most_common()
            ],
            "topCouncils": sorted(
                [
                    {
                        "name": council,
                        "ships": council_ships[council],
                        "youth": youth,
                    }
                    for council, youth in council_youth.items()
                ],
                key=lambda item: item["youth"],
                reverse=True,
            )[:15],
            "councils": top_counts(ships, "council", 20),
        },
        "ships": ships,
    }
    return payload


def main():
    capture_path = latest_capture()
    rows = load_rows(capture_path)
    payload = build_payload(rows, capture_path)
    JSON_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {JSON_PATH} with {len(rows)} ships from {capture_path.name}")


if __name__ == "__main__":
    main()
